"""다운로드된 OLAP Excel 파일 파싱 — 3개 리포트별 데이터 추출.

# ── 파싱 튜닝 가이드 ─────────────────────────────────────────────────
# OLAP 에서 다운로드한 Excel 파일의 실제 컬럼명/구조를 확인하려면:
#   python -m src.excel_parser <파일경로>
# 위 명령으로 헤더와 샘플 행이 출력됩니다.
#
# 컬럼명이 다를 경우 아래 COLUMN_MAP_* 상수를 수정하세요.
# ────────────────────────────────────────────────────────────────────
"""

from __future__ import annotations

import logging
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl

log = logging.getLogger(__name__)


# ── .xls 어댑터 (xlrd 래퍼) ───────────────────────────────────────────

class _WorksheetAdapter:
    """xlrd 시트를 openpyxl 워크시트처럼 다루기 위한 경량 어댑터."""

    def __init__(self, title: str, data: list[tuple]):
        self.title = title
        self._data = data
        self.max_row = len(data)
        self.max_column = max((len(r) for r in data), default=0)

    def iter_rows(self, values_only: bool = True, min_row: int = 1):
        for row in self._data[min_row - 1:]:
            yield row


def _load_xls_sheet(path: Path) -> _WorksheetAdapter:
    """XLS 파일 로드 — 실제 바이너리 XLS 또는 HTML 위장 XLS 모두 처리."""
    raw = path.read_bytes()
    if raw[:2] != b"\xd0\xcf":  # OLE compound document magic bytes 아님 → HTML
        return _load_html_table_sheet(path, raw)

    import xlrd
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    rows: list[tuple] = []
    for r in range(ws.nrows):
        row: list = []
        for c in range(ws.ncols):
            cell = ws.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                val = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                row.append(val.date() if val.hour == 0 and val.minute == 0 else val)
            elif cell.ctype == xlrd.XL_CELL_NUMBER:
                n = cell.value
                row.append(int(n) if n == int(n) else n)
            elif cell.ctype == xlrd.XL_CELL_TEXT:
                row.append(cell.value)
            else:
                row.append(None if cell.ctype in (0, 6) else cell.value)
        rows.append(tuple(row))
    log.debug(f"  XLS 로드: {path.name} / 시트={ws.name} / {ws.nrows}행×{ws.ncols}열")
    return _WorksheetAdapter(ws.name, rows)


def _load_html_table_sheet(path: Path, raw: bytes) -> _WorksheetAdapter:
    """HTML 형식으로 저장된 Excel 파일(OLAP 다운로드)을 파싱합니다."""
    from html.parser import HTMLParser

    class _TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows: list[list[str]] = []
            self._cur_row: list[str] | None = None
            self._cur_cell: str | None = None
            self._in_cell = False

        def handle_starttag(self, tag, attrs):
            if tag == "tr":
                self._cur_row = []
            elif tag in ("td", "th") and self._cur_row is not None:
                self._cur_cell = ""
                self._in_cell = True

        def handle_endtag(self, tag):
            if tag in ("td", "th") and self._in_cell:
                self._cur_row.append(self._cur_cell.strip())  # type: ignore[union-attr]
                self._cur_cell = None
                self._in_cell = False
            elif tag == "tr" and self._cur_row is not None:
                if any(self._cur_row):
                    self.rows.append(self._cur_row)
                self._cur_row = None

        def handle_data(self, data):
            if self._in_cell and self._cur_cell is not None:
                self._cur_cell += data

    # charset 우선 감지 — meta 태그 선언 기준
    peek = raw[:500].decode("latin-1").lower()
    if "charset=utf-8" in peek or 'charset="utf-8"' in peek:
        html = raw.decode("utf-8", errors="replace")
    else:
        html = raw.decode("euc-kr", errors="replace")

    parser = _TableParser()
    parser.feed(html)

    def _coerce(v: str):
        v = v.replace(",", "").strip()
        if not v or v in ("\xa0", "."):  # SAS missing value "." → None
            return None
        # 날짜 형식 시도 (유니코드 대시 정규화)
        from datetime import datetime as _dt
        vd = _normalize_dashes(v)
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
            try:
                return _dt.strptime(vd, fmt).date()
            except ValueError:
                pass
        # 숫자 시도
        try:
            f = float(v)
            return int(f) if f == int(f) else f
        except ValueError:
            return v

    # 단일 셀 행(설명 테이블) 제거 → 다중 컬럼 행만 유지
    rows = [tuple(_coerce(c) for c in row) for row in parser.rows if len(row) > 1]
    log.debug(f"  HTML 테이블 로드: {path.name} / {len(rows)}행")
    return _WorksheetAdapter(path.stem, rows)

# ── 리포트 1: [HPC] 일별 회원관리지표 ──────────────────────────────
# Excel 컬럼명 → 대상 시트 컬럼명 매핑 (왼쪽: OLAP Excel, 오른쪽: Google Sheets)
MEMBER_COL_MAP: dict[str, str] = {
    "신규가입회원수":     "신규회원수",
    "해피앱 로그인 회원수": "해피앱 로그인수",
    "해피앱 DAU":        "해피앱 DAU",
    "해피오더 DAU":      "해피오더 DAU",
}
# Excel 에서 날짜가 들어있는 컬럼명 (행 필터링용)
MEMBER_DATE_COL = "일자"  # TODO: 실제 OLAP Excel 헤더 확인 후 수정

# ── 리포트 2: [HPC] 채널별 적립, 사용건수 현황 ─────────────────────
# 채널명 식별값 — 컬럼 헤더(wide 포맷) 또는 행 값(tall 포맷) 모두 지원
CHANNEL_TARGET_LABEL = "HPCAPP"
# tall 포맷: 채널명 컬럼명 (HPCAPP 이 행 값으로 존재할 때)
CHANNEL_KEY_COL = "채널"
# tall 포맷: 건수 컬럼명
CHANNEL_VALUE_COL = "제시건수"

# ── 리포트 3: [HPC, POS] HPC 일마감(브랜드) ────────────────────────
# 브랜드 식별자 — 컬럼 헤더로 존재 (유저 확인)
CLOSING_BRAND_LABEL = "0002. SPC전사(3사)"
# 메트릭 행 레이블이 들어있는 컬럼명 (행 인덱스 열)
CLOSING_ROW_KEY_COL = 0  # 첫 번째 열 (0-indexed) — TODO: 실제 구조 확인 후 수정
# 메트릭 행 레이블 → 대상 시트 컬럼명 매핑
CLOSING_ROW_MAP: dict[str, str] = {
    "POS 총매출액":  "POS 총매출액",
    "POS 영수증건수": "POS 영수증건수",
    "POS 거래점포수": "POS 거래점포수",
    "HPC 매출액":    "HPC 매출액",
    "HPC 거래점포수": "HPC 거래점포수",
    "HPC 총적립액":  "HPC 총적립액",
    "HPC 적립건수":  "HPC 적립건수",
    "객단가":        "객단가",
    "HPC 총사용액":  "HPC 총사용액",
    "HPC 사용건수":  "HPC 사용건수",
}


# ── 공통 유틸 ─────────────────────────────────────────────────────────

def _load_first_sheet(path: Path):
    if path.suffix.lower() == ".xls":
        return _load_xls_sheet(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    log.debug(f"  엑셀 로드: {path.name} / 시트={ws.title} / {ws.max_row}행×{ws.max_column}열")
    return ws


def _header_row(ws) -> list[str]:
    """첫 번째 비어있지 않은 행을 헤더로 반환."""
    for row in ws.iter_rows(values_only=True):
        cleaned = [str(c).strip() if c is not None else "" for c in row]
        if any(cleaned):
            return cleaned
    return []


def _normalize_dashes(s: str) -> str:
    """유니코드 대시(−,–,—,‐)를 ASCII '-' 로 정규화. WRS 일마감 날짜가 U+2212 를 씀."""
    if not s:
        return s
    return (
        s.replace("−", "-")  # MINUS SIGN
         .replace("–", "-")  # EN DASH
         .replace("—", "-")  # EM DASH
         .replace("‐", "-")  # HYPHEN
    )


def _parse_date_label(s: str) -> date | None:
    """'2026-06-14' 등 날짜 라벨 문자열을 date 로 (대시 정규화 후)."""
    from datetime import datetime as _dt
    s = _normalize_dashes(str(s)).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return _dt.strptime(s, fmt).date()
        except ValueError:
            continue
    import re as _re
    m = _re.match(r"\s*(\d{4})\D+(\d{1,2})\D+(\d{1,2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _to_num(val: Any) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _dates_match(cell_val: Any, target: date) -> bool:
    """셀 값(날짜 직렬/문자열/datetime)이 target date 와 일치하는지 확인."""
    if cell_val is None:
        return False
    from datetime import datetime as dt
    if isinstance(cell_val, (dt,)):
        return cell_val.date() == target
    if isinstance(cell_val, date):
        return cell_val == target
    # 문자열 형식 시도 (유니코드 대시 정규화)
    s = _normalize_dashes(str(cell_val)).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d/%Y"):
        try:
            return dt.strptime(s, fmt).date() == target
        except ValueError:
            continue
    return False


def _safe_print(label: str, value) -> None:
    try:
        print(f"{label}: {value}")
    except UnicodeEncodeError:
        encoded = str(value).encode("ascii", errors="replace").decode("ascii")
        print(f"{label}: {encoded}")


def _debug_dump(path: Path) -> None:
    """헤더 + 상위 5행 출력 (파싱 튜닝용)."""
    ws = _load_first_sheet(path)
    headers = _header_row(ws)
    _safe_print(f"\n=== {path.name} ===\n헤더", headers)
    count = 0
    for row in ws.iter_rows(values_only=True):
        if count >= 5:
            break
        if any(c is not None for c in row):
            _safe_print("행", row[:20])
            count += 1


# ── 리포트 1 파서 ─────────────────────────────────────────────────────

def parse_member_metrics(path: Path, target_date: date) -> dict[str, Any]:
    """[HPC] 일별 회원관리지표 Excel → 어제 날짜 행 값 추출."""
    log.info(f"[PARSE] 회원지표 파싱: {path.name}")
    ws = _load_first_sheet(path)
    headers = _header_row(ws)

    if not headers:
        raise ValueError(f"헤더를 찾을 수 없습니다: {path}")

    # 날짜 컬럼 인덱스
    date_idx: int | None = None
    for i, h in enumerate(headers):
        if h == MEMBER_DATE_COL or "일자" in h or "날짜" in h or "date" in h.lower():
            date_idx = i
            break

    if date_idx is None:
        log.warning(
            f"  날짜 컬럼 '{MEMBER_DATE_COL}' 미발견. 헤더: {headers[:15]}\n"
            "  MEMBER_DATE_COL 상수를 실제 컬럼명으로 수정하세요."
        )
        # fallback: 첫 번째 데이터 행의 값으로 매핑
        date_idx = 0

    # 대상 행 탐색
    target_row: list | None = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if _dates_match(row[date_idx], target_date):
            target_row = list(row)
            break

    if target_row is None:
        raise ValueError(
            f"날짜 {target_date} 에 해당하는 행이 없습니다 (파일: {path.name}). "
            "데이터가 아직 준비되지 않았거나, MEMBER_DATE_COL 컬럼명을 확인하세요."
        )

    result: dict[str, Any] = {}
    for src_col, dst_col in MEMBER_COL_MAP.items():
        if src_col in headers:
            idx = headers.index(src_col)
            result[dst_col] = _to_num(target_row[idx])
        else:
            log.warning(f"  컬럼 '{src_col}' 미발견 (헤더: {headers[:15]})")
            result[dst_col] = None

    log.info(f"  추출 결과: {result}")
    return result


# ── 리포트 2 파서 ─────────────────────────────────────────────────────

def parse_channel_metrics(path: Path, target_date: date) -> dict[str, Any]:
    """[HPC] 채널별 적립, 사용건수 현황 → HPCAPP 제시건수 추출.

    Wide 포맷 (HPCAPP 이 컬럼 헤더):
      헤더: [일자, 브랜드, HPC전체, HPCAPP, ...]
      데이터: [2026-05-29, SPC전사(3사), 237189, 95467, ...]

    Tall 포맷 (HPCAPP 이 행 값):
      헤더: [일자, 채널, 제시건수, ...]
      데이터: [2026-05-29, HPCAPP, 95467, ...]
    """
    log.info(f"[PARSE] 채널별 지표 파싱: {path.name}")
    ws = _load_first_sheet(path)
    headers = _header_row(ws)

    if not headers:
        raise ValueError(f"헤더를 찾을 수 없습니다: {path}")

    # ── Wide 포맷 감지: HPCAPP 이 헤더에 직접 있는 경우 ───────────────
    if CHANNEL_TARGET_LABEL in headers:
        val_idx = headers.index(CHANNEL_TARGET_LABEL)
        date_idx: int | None = None
        for i, h in enumerate(headers):
            if "일자" in h or "날짜" in h or "date" in h.lower():
                date_idx = i
                break
        for row in ws.iter_rows(min_row=2, values_only=True):
            if date_idx is not None and not _dates_match(row[date_idx], target_date):
                continue
            val = _to_num(row[val_idx])
            result = {"APP 제시건수": val}
            log.info(f"  추출 결과 (wide 포맷): {result}")
            return result
        raise ValueError(
            f"Wide 포맷에서 날짜 {target_date} 에 해당하는 행 없음 (파일: {path.name})"
        )

    # ── Tall 포맷: HPCAPP 이 행 값으로 존재 ─────────────────────────
    date_idx = None
    for i, h in enumerate(headers):
        if "일자" in h or "날짜" in h or "date" in h.lower():
            date_idx = i
            break

    key_idx: int | None = None
    for i, h in enumerate(headers):
        if h == CHANNEL_KEY_COL or "채널" in h or "channel" in h.lower():
            key_idx = i
            break

    val_idx2: int | None = None
    for i, h in enumerate(headers):
        if h == CHANNEL_VALUE_COL or "제시건수" in h or "건수" in h:
            val_idx2 = i
            break

    if key_idx is None or val_idx2 is None:
        log.warning(
            f"  채널 컬럼 미발견. 헤더: {headers[:15]}\n"
            "  CHANNEL_KEY_COL / CHANNEL_VALUE_COL 을 수정하세요."
        )
        raise ValueError(f"채널 컬럼 미발견: {path.name}")

    for row in ws.iter_rows(min_row=2, values_only=True):
        if date_idx is not None and not _dates_match(row[date_idx], target_date):
            continue
        cell_key = str(row[key_idx]).strip() if row[key_idx] else ""
        if CHANNEL_TARGET_LABEL in cell_key:
            val = _to_num(row[val_idx2])
            result = {"APP 제시건수": val}
            log.info(f"  추출 결과 (tall 포맷): {result}")
            return result

    raise ValueError(
        f"'{CHANNEL_TARGET_LABEL}' 행이 없습니다 (파일: {path.name}). "
        "CHANNEL_TARGET_LABEL / CHANNEL_KEY_COL 을 확인하세요."
    )


# ── 리포트 3 파서 ─────────────────────────────────────────────────────

# 일마감 메트릭 이름 정규화 매핑 (공백 제거 헤더 → 대상 시트 컬럼명)
CLOSING_METRIC_NORM: dict[str, str] = {
    "POS총매출액": "POS 총매출액",
    "POS영수증건수": "POS 영수증건수",
    "POS거래점포수": "POS 거래점포수",
    "HPC매출액": "HPC 매출액",
    "HPC거래점포수": "HPC 거래점포수",
    "HPC총적립액": "HPC 총적립액",
    "HPC적립건수": "HPC 적립건수",
    "객단가": "객단가",
    "HPC총사용액": "HPC 총사용액",
    "HPC사용건수": "HPC 사용건수",
}


class ClosingDateNotFoundError(ValueError):
    """일마감 리포트에 target_date 열그룹이 없을 때 (데이터 미준비 가능)."""


def parse_closing_report(path: Path, target_date: date) -> dict[str, Any]:
    """[HPC, POS] HPC 일마감(브랜드) → '0002. SPC전사(3사)' 행, **target_date 열그룹** 추출.

    일마감 리포트는 [대상일, 전일] 두 날짜를 나란히 보여준다(바깥 열차원 c0=날짜).
    반드시 target_date 에 해당하는 c0 그룹의 값만 추출해야 전일 데이터 혼입을 막는다.

    파일 형식에 따라 자동 분기:
    - .html : SAS WRS iFrame HTML (브랜드=행, 날짜=바깥열 c0, 메트릭=안쪽열 c1)
    - .xlsx/.xls : OLAP Excel 다운로드 (구 방식)
    """
    if path.suffix.lower() == ".html":
        return _parse_closing_report_html(path, target_date)
    return _parse_closing_report_excel(path)


def _parse_closing_report_html(path: Path, target_date: date) -> dict[str, Any]:
    """SAS WRS HTML에서 0002.SPC전사(3사) 행 × target_date 열그룹 데이터 추출.

    WRS OLAP 테이블 구조 (셀의 headers 속성이 그룹을 명시):
    - 행 헤더 r1_N: 브랜드명 (0002.SPC전사(3사) = r1_0)
    - 바깥 컬럼 c0_N: 날짜 (예 c0_0=대상일, c0_10=전일) — <span id="..._c0_N_..label">날짜
    - 안쪽 컬럼 c1_N: 메트릭명 — <span id="..._c1_N_..label">메트릭
    - 데이터 <td headers="... r1_0 ... c0_K ... c1_M">: 해당 (브랜드,날짜,메트릭) 값
    """
    log.info(f"[PARSE] WRS 일마감 HTML 파싱: {path.name} (대상일 {target_date})")
    html = path.read_text(encoding="utf-8", errors="replace")

    # 테이블 ID prefix 탐색 (otvc{숫자}_otv)
    m = re.search(r'id="(otvc\w+_otv)_r1_0', html)
    if not m:
        raise ValueError(f"WRS 테이블 prefix 미발견: {path}")
    pfx = m.group(1)

    # ── 바깥 컬럼(c0=날짜) 라벨 → date 매핑 ──────────────────────────
    c0_pat = re.compile(
        r'id="' + re.escape(pfx) + r'_c0_(\d+)_[^"]*label"[^>]*>([^<]+)</span>'
    )
    c0_dates: dict[int, date] = {}
    for hm in c0_pat.finditer(html):
        d = _parse_date_label(hm.group(2))
        if d is not None:
            c0_dates[int(hm.group(1))] = d

    target_c0 = [idx for idx, d in c0_dates.items() if d == target_date]
    if not target_c0:
        raise ClosingDateNotFoundError(
            f"일마감 리포트에 {target_date} 열그룹 없음 "
            f"(존재 날짜: {sorted(set(c0_dates.values()))}) — 데이터 미준비 가능"
        )
    c0_idx = target_c0[0]
    c0_tok = f"{pfx}_c0_{c0_idx}"
    brand_tok = f"{pfx}_r1_0"  # 0002.SPC전사(3사)

    # ── 안쪽 컬럼(c1=메트릭) 라벨 매핑 ───────────────────────────────
    c1_pat = re.compile(
        r'id="' + re.escape(pfx) + r'_c1_(\d+)_[^"]*label"[^>]*>([^<]+)</span>'
    )
    c1_labels: dict[int, str] = {}
    for hm in c1_pat.finditer(html):
        c1_labels[int(hm.group(1))] = hm.group(2).strip().replace(" ", "").replace("\xa0", "")

    # ── 데이터 셀: headers 에 브랜드 r1_0 AND 대상 날짜 c0_idx 토큰을 모두 포함 ──
    cell_pat = re.compile(
        r'<td\s+id="' + re.escape(pfx) + r'_[^"]*_cr"\s+headers="([^"]*)"[^>]*>([^<]*)</td>',
        re.IGNORECASE,
    )
    c1_tok_pat = re.compile(re.escape(pfx) + r"_c1_(\d+)$")

    result: dict[str, Any] = {}
    for hm in cell_pat.finditer(html):
        tokens = hm.group(1).split()
        if brand_tok not in tokens or c0_tok not in tokens:
            continue  # 다른 브랜드 또는 다른 날짜 그룹 → 제외
        c1_idx = None
        for t in tokens:
            cm = c1_tok_pat.match(t)
            if cm:
                c1_idx = int(cm.group(1))
                break
        if c1_idx is None:
            continue
        hdr_raw = c1_labels.get(c1_idx, "")
        for src, dst in CLOSING_METRIC_NORM.items():
            if src in hdr_raw or dst.replace(" ", "") in hdr_raw:
                if dst not in result:  # 같은 그룹 내 첫 매칭 유지
                    result[dst] = _to_num(hm.group(2).strip())
                break

    missing = [v for v in CLOSING_METRIC_NORM.values() if v not in result]
    if missing:
        log.warning(f"  미추출 항목: {missing} (c0_idx={c0_idx}, 헤더 매핑 확인)")

    log.info(f"  추출 결과 ({len(result)}개, 날짜 {target_date}): {result}")
    return result


def _parse_closing_report_excel(path: Path) -> dict[str, Any]:
    """Excel 파일에서 0002. SPC전사(3사) 컬럼 추출 (구 방식)."""
    log.info(f"[PARSE] 일마감 리포트 파싱: {path.name}")
    ws = _load_first_sheet(path)
    headers = _header_row(ws)

    if not headers:
        raise ValueError(f"헤더를 찾을 수 없습니다: {path}")

    # "0002. SPC전사(3사)" 컬럼 인덱스 탐색
    brand_idx: int | None = None
    for i, h in enumerate(headers):
        if CLOSING_BRAND_LABEL in h:
            brand_idx = i
            break

    if brand_idx is None:
        # fallback: 부분 문자열 "SPC전사" 로 재탐색
        for i, h in enumerate(headers):
            if "SPC전사" in h or "전사(3사)" in h:
                brand_idx = i
                log.warning(f"  '{CLOSING_BRAND_LABEL}' 대신 '{headers[i]}' 컬럼 사용")
                break

    if brand_idx is None:
        log.warning(
            f"  브랜드 컬럼 '{CLOSING_BRAND_LABEL}' 미발견. 헤더: {headers[:20]}\n"
            "  CLOSING_BRAND_LABEL 을 실제 컬럼명으로 수정하세요."
        )
        raise ValueError(f"브랜드 컬럼 미발견: {path.name}")

    # 메트릭 행 추출
    result: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_key = str(row[CLOSING_ROW_KEY_COL]).strip() if row[CLOSING_ROW_KEY_COL] else ""
        for src_label, dst_col in CLOSING_ROW_MAP.items():
            if src_label in row_key:
                result[dst_col] = _to_num(row[brand_idx])
                break

    missing = [v for v in CLOSING_ROW_MAP.values() if v not in result]
    if missing:
        log.warning(
            f"  미추출 항목: {missing}\n"
            "  CLOSING_ROW_MAP 의 행 레이블을 실제 Excel 값으로 수정하세요."
        )

    log.info(f"  추출 결과: {result}")
    return result


# ── VISUAL REPORT MAU 파서 ───────────────────────────────────────────

def parse_mau_excel(path: Path) -> int:
    """VISUAL REPORT MAU 당월 Excel → A2 셀 정수값 반환."""
    log.info(f"[PARSE] MAU Excel 파싱: {path.name}")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active or wb[wb.sheetnames[0]]
    val = ws.cell(row=2, column=1).value
    if val is None:
        raise ValueError(f"MAU Excel A2 셀이 비어 있습니다 (파일: {path.name})")
    result = int(float(str(val).replace(",", "").strip()))
    log.info(f"  MAU 당월: {result:,}")
    return result


# ── CLI 디버그 모드 ───────────────────────────────────────────────────

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format="%(message)s")
    if len(sys.argv) < 2:
        print("사용법: python -m src.excel_parser <파일경로>")
        sys.exit(1)
    _debug_dump(Path(sys.argv[1]))
